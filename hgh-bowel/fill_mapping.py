# -*- coding: utf-8 -*-
"""Merge work/result.csv back into the mapping workbook.

Safety rules baked in:
  - the original workbook is backed up before any write
  - a cell that already holds a value is never overwritten unless --overwrite
  - 種族 (F) and 就醫來源 (G) are never touched - they are filled by hand
  - rows whose accession has no result are left alone and reported

Console output is cp950-safe: no arrows, no check marks.

Usage:
    python fill_mapping.py --xlsx "<path>" --dry-run
    python fill_mapping.py --xlsx "<path>"
"""
import argparse
import csv
import os
import shutil
import sys
import time

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
WORK_DIR = os.path.join(HERE, "work")
RESULT_CSV = os.path.join(WORK_DIR, "result.csv")

SHEET = "mapping"
COL_ACCESSION = 2

# result.csv column -> worksheet column index.
# F (種族) and G (就醫來源) are deliberately absent: filled by hand, not from DICOM.
TARGET_COLUMNS = {
    "拍攝年": 3,    # C
    "年齡": 4,      # D
    "性別": 5,      # E
    "解析度": 8,    # H
    "CT廠牌": 9,    # I
    "CT型號": 10,   # J
}


def read_results():
    if not os.path.exists(RESULT_CSV):
        sys.exit("ERROR: 找不到 %s，請先擷取至少一筆" % RESULT_CSV)
    results = {}
    dupes = []
    with open(RESULT_CSV, "r", encoding="utf-8-sig", newline="") as fh:
        for rec in csv.DictReader(fh):
            acc = (rec.get("AccessionNumber") or "").strip()
            if not acc:
                continue
            if acc in results:
                dupes.append(acc)
            results[acc] = rec  # a later re-capture supersedes an earlier one
    return results, dupes


def backup(path):
    stamp = time.strftime("%Y%m%d-%H%M%S")
    base, ext = os.path.splitext(path)
    dst = "%s.bak-%s%s" % (base, stamp, ext)
    shutil.copy2(path, dst)
    return dst


def main():
    ap = argparse.ArgumentParser(description="Merge results into the mapping workbook")
    ap.add_argument("--xlsx", required=True, help="path to the mapping workbook")
    ap.add_argument("--dry-run", action="store_true", help="report only, write nothing")
    ap.add_argument("--overwrite", action="store_true",
                    help="also replace cells that already hold a value")
    args = ap.parse_args()

    results, dupes = read_results()
    if dupes:
        print("注意：result.csv 有 %d 筆重複單號，採用最後一筆" % len(dupes))

    wb = openpyxl.load_workbook(args.xlsx)
    if SHEET not in wb.sheetnames:
        sys.exit("ERROR: 找不到工作表 %r" % SHEET)
    ws = wb[SHEET]

    written = 0
    skipped_filled = 0
    rows_touched = 0
    unmatched_rows = []
    matched_accessions = set()
    blank_values = []

    for r in range(2, ws.max_row + 1):
        acc_cell = ws.cell(r, COL_ACCESSION).value
        if acc_cell is None:
            continue
        acc = str(acc_cell).strip()
        rec = results.get(acc)
        if not rec:
            unmatched_rows.append(acc)
            continue
        matched_accessions.add(acc)
        touched = False
        for key, col in TARGET_COLUMNS.items():
            value = (rec.get(key) or "").strip()
            if not value:
                blank_values.append((acc, key))
                continue
            cell = ws.cell(r, col)
            if cell.value not in (None, "") and not args.overwrite:
                skipped_filled += 1
                continue
            if not args.dry_run:
                # 年齡/拍攝年/性別 stay numeric so the KMU side can aggregate them.
                cell.value = int(value) if key in ("拍攝年", "年齡", "性別") and value.isdigit() else value
            written += 1
            touched = True
        if touched:
            rows_touched += 1

    print("result.csv 可用筆數    %d" % len(results))
    print("對到的 mapping 列      %d" % len(matched_accessions))
    print("寫入儲存格             %d（涉及 %d 列）" % (written, rows_touched))
    if skipped_filled:
        print("略過已有值的儲存格     %d（要覆蓋請加 --overwrite）" % skipped_filled)
    if blank_values:
        print("空值未寫入             %d 欄，例如 %s"
              % (len(blank_values), ", ".join("%s/%s" % b for b in blank_values[:3])))
    if unmatched_rows:
        print("尚無擷取結果的列       %d" % len(unmatched_rows))
    orphan = set(results) - matched_accessions
    if orphan:
        print("注意：result.csv 有 %d 筆單號在 mapping 表找不到：%s"
              % (len(orphan), ", ".join(sorted(orphan)[:5])))

    if args.dry_run:
        print("")
        print("(dry-run，未寫檔)")
        return 0

    if written:
        bak = backup(args.xlsx)
        print("")
        print("已備份 %s" % os.path.basename(bak))
        wb.save(args.xlsx)
        print("已寫回 %s" % os.path.basename(args.xlsx))
    else:
        print("")
        print("沒有需要寫入的內容，未動原檔")
    return 0


if __name__ == "__main__":
    sys.exit(main())
