# -*- coding: utf-8 -*-
"""Export the case/accession queue from the mapping workbook, and show progress.

The workbook itself is never modified here - this is a read-only step.
Console output is cp950-safe: no arrows, no check marks.

Usage:
    python export_queue.py --xlsx "<path to mapping xlsx>"
    python export_queue.py --xlsx "<...>" --status     # show what is still outstanding
    python export_queue.py --xlsx "<...>" --next 10    # print the next 10 to do
"""
import argparse
import csv
import os
import sys

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
WORK_DIR = os.path.join(HERE, "work")
QUEUE_CSV = os.path.join(WORK_DIR, "queue.csv")
RESULT_CSV = os.path.join(WORK_DIR, "result.csv")

SHEET = "mapping"
COL_CASE = 1        # A 資料夾名稱
COL_ACCESSION = 2   # B 申請單號


def read_queue_from_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit("ERROR: 找不到工作表 %r，實際有：%s" % (SHEET, wb.sheetnames))
    ws = wb[SHEET]
    rows = []
    for r in range(2, ws.max_row + 1):
        case = ws.cell(r, COL_CASE).value
        acc = ws.cell(r, COL_ACCESSION).value
        if case is None and acc is None:
            continue
        rows.append({
            "row": r,
            "case": str(case).strip() if case is not None else "",
            "accession": str(acc).strip() if acc is not None else "",
        })
    return rows


def read_done_accessions():
    if not os.path.exists(RESULT_CSV):
        return set()
    done = set()
    with open(RESULT_CSV, "r", encoding="utf-8-sig", newline="") as fh:
        for rec in csv.DictReader(fh):
            acc = (rec.get("AccessionNumber") or "").strip()
            if acc:
                done.add(acc)
    return done


def write_queue(rows):
    os.makedirs(WORK_DIR, exist_ok=True)
    with open(QUEUE_CSV, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["row", "case", "accession"])
        writer.writeheader()
        writer.writerows(rows)


def main():
    ap = argparse.ArgumentParser(description="Export the case/accession queue")
    ap.add_argument("--xlsx", required=True, help="path to the mapping workbook")
    ap.add_argument("--status", action="store_true", help="show progress only")
    ap.add_argument("--next", type=int, metavar="N", help="print the next N outstanding cases")
    args = ap.parse_args()

    rows = read_queue_from_xlsx(args.xlsx)
    done = read_done_accessions()
    outstanding = [r for r in rows if r["accession"] not in done]

    dupes = len(rows) - len({r["accession"] for r in rows if r["accession"]})
    blanks = [r for r in rows if not r["accession"]]

    print("總筆數      %d" % len(rows))
    print("已擷取      %d" % (len(rows) - len(outstanding)))
    print("待處理      %d" % len(outstanding))
    if dupes:
        print("注意：申請單號重複 %d 筆" % dupes)
    if blanks:
        print("注意：申請單號空白 %d 筆（%s）"
              % (len(blanks), ", ".join(b["case"] for b in blanks[:5])))

    if args.next:
        print("")
        print("接下來 %d 筆：" % min(args.next, len(outstanding)))
        for r in outstanding[:args.next]:
            print("  %-10s %s" % (r["case"], r["accession"]))

    if not args.status and not args.next:
        write_queue(rows)
        print("")
        print("已寫出 %s" % QUEUE_CSV)

    return 0


if __name__ == "__main__":
    sys.exit(main())
