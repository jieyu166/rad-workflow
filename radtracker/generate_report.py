#!/usr/bin/env python3
"""
generate_report.py — 產出放射科週報資料

Usage (CSV mode - recommended):
    python generate_report.py --csv 202602.csv --input week_input.yaml -o output/weekly_report.json
    python generate_report.py --csv 202602.csv --yk 202602YK.csv --input week_input.yaml

Usage (Legacy XLSX mode):
    python generate_report.py --xlsx tracker.xlsx --input week_input.md -o output/weekly_report.json
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

# ── Baselines (from claude.md) ──────────────────────────────────────────────
BASELINES = {
    'X光(普通)': 50, 'X光(急打)': 35, 'X光(中等)': 30, 'X光(困難)': 20,
    'CT(Chest/Abd)': 3.5, 'CT(Brain)': 7, 'CT(Neck/CTA)': 3.5, 'CT(Brain-C)': 3,
    'US(一般)': 18, 'US(困難)': 4,
    'Mammo': 22, 'MR': 2, 'IVP': 12, 'BMD': 60,
}

# Modality name mapping (CSV uses English, XLSX uses Chinese for XR)
CSV_TO_DISPLAY = {'XR': 'X光', 'CT': 'CT', 'US': 'US', 'Mammo': 'Mammo',
                  'MR': 'MR', 'BMD': 'BMD', 'IVP': 'IVP', 'Other': '其他'}
XLSX_MOD_MAP = {'xr': 'X光', 'ct': 'CT', 'us': 'US', 'mm': 'Mammo', 'mr': 'MR',
                'bmd': 'BMD', 'other': '其他', 'ivp': 'IVP'}

TRACKED_MODS = ['X光', 'CT', 'US', 'Mammo']
ALL_MODS = ['X光', 'CT', 'US', 'Mammo', 'MR', 'BMD', 'IVP']
DAY_ORDER = ['一', '二', '三', '四', '五', '六', '日']


# ══════════════════════════════════════════════════════════════════════════
#  CSV MODE
# ══════════════════════════════════════════════════════════════════════════

def parse_yaml_input(path: str) -> dict:
    """Parse week_input.yaml for remaining counts and metadata.
    Uses a simple parser to avoid pyyaml dependency."""
    text = Path(path).read_text(encoding='utf-8')
    info = {
        'start': {}, 'end': {}, 'mid': {}, 'week': '', 'date_range': '',
        'duty_day': '', 'notes': [], 'study_hr': 0,
    }

    # Try pyyaml first
    try:
        import yaml
        data = yaml.safe_load(text)
        info['week'] = data.get('week', '')
        info['date_range'] = data.get('date_range', '')
        info['duty_day'] = data.get('duty_day', '')
        info['notes'] = data.get('notes', [])

        remaining = data.get('remaining', {})
        start = remaining.get('start', {})
        end = remaining.get('end', {})
        mid = remaining.get('mid', {})

        # Map to display names
        mod_map = {'XR': 'X光', 'CT': 'CT', 'US': 'US', 'Mammo': 'Mammo',
                   'IVP': 'IVP', 'BMD': 'BMD', 'MR': 'MR'}
        for k, v in start.items():
            display = mod_map.get(k, k)
            info['start'][display] = int(v)
        for k, v in end.items():
            display = mod_map.get(k, k)
            info['end'][display] = int(v)
        for k, v in mid.items():
            display = mod_map.get(k, k)
            info['mid'][display] = int(v)

        # Study hours
        study = data.get('study', [])
        if study:
            info['study_hr'] = sum(s.get('hours', 0) for s in study if isinstance(s, dict))

        return info
    except ImportError:
        pass

    # Fallback: simple regex parsing
    for line in text.split('\n'):
        line = line.strip()
        if line.startswith('week:'):
            info['week'] = line.split(':', 1)[1].strip().strip('"\'')
        elif line.startswith('date_range:'):
            info['date_range'] = line.split(':', 1)[1].strip().strip('"\'')
        elif line.startswith('duty_day:'):
            info['duty_day'] = line.split(':', 1)[1].strip().strip('"\'')

    # Parse remaining block with regex
    # Look for patterns like {XR: 617, CT: 16, US: 9, Mammo: 76}
    mod_map = {'XR': 'X光', 'CT': 'CT', 'US': 'US', 'Mammo': 'Mammo',
               'IVP': 'IVP', 'BMD': 'BMD', 'MR': 'MR'}

    for key in ('start', 'end', 'mid'):
        match = re.search(key + r':\s*\{([^}]+)\}', text)
        if match:
            for pair in match.group(1).split(','):
                k, v = pair.split(':')
                k = k.strip()
                display = mod_map.get(k, k)
                info[key][display] = int(v.strip())

    return info


def load_csv_data(csv_paths: list[str], yk_path: str | None,
                  week_str: str, reporter_id: str) -> dict:
    """Load and process CSV data using parse_csv module."""
    from parse_csv import process_csvs

    all_paths = list(csv_paths)
    if yk_path:
        all_paths.append(yk_path)

    return process_csvs(all_paths, week_str, reporter_id)


def csv_to_report_data(csv_data: dict, week_input: dict) -> dict:
    """Convert parse_csv output to the standard report data format."""
    # Map CSV modality names to display names
    daily = {}
    for wd in DAY_ORDER:
        d = csv_data['daily'][wd]
        day_data = {
            'total': d['cases'],
            'report_min': round(d['active_hr'] * 60, 0),  # convert hr to min
            'active_hr': d['active_hr'],
            'span_hr': d['span_hr'],
            'sessions': d['sessions'],
            'points': d['points'],
            'clinical_min': 0,  # CSV doesn't track clinical time
        }
        for m_csv, m_disp in CSV_TO_DISPLAY.items():
            if d.get(m_csv, 0) > 0:
                day_data[m_disp] = d[m_csv]
        if 'xr_sub' in d:
            day_data['xr_sub'] = d['xr_sub']
        if 'ct_sub' in d:
            day_data['ct_sub'] = d['ct_sub']
        if 'us_sub' in d:
            day_data['us_sub'] = d['us_sub']
        daily[wd] = day_data

    # Totals
    totals_mods = {}
    for m_csv, cnt in csv_data['totals']['mods'].items():
        m_disp = CSV_TO_DISPLAY.get(m_csv, m_csv)
        totals_mods[m_disp] = cnt

    totals = {
        'mods': totals_mods,
        'report_min': round(csv_data['totals']['active_hr'] * 60, 0),
        'active_hr': csv_data['totals']['active_hr'],
        'clinical_min': 0,
        'misc_min': 0,
        'grand_total': csv_data['totals']['grand_total'],
        'points': csv_data['totals']['points'],
    }

    # Tracking (from user-provided remaining counts)
    tracking = compute_tracking(week_input, totals)

    # Efficiency (sub-category level, estimated from daily)
    efficiency = {}
    for key, data in csv_data.get('sub_counts', {}).items():
        # Map CSV key to display key
        display_key = key.replace('XR(', 'X光(')
        count = data['count']
        baseline = BASELINES.get(display_key)
        efficiency[display_key] = {
            'count': count,
            'mins': None,  # no per-sub timing from CSV
            'actual_rate': None,
            'baseline': baseline,
            'diff_pct': None,
            'estimated': True,
        }

    # Daily efficiency (estimated)
    efficiency_daily = {}
    for wd, eff in csv_data.get('efficiency_daily', {}).items():
        efficiency_daily[wd] = eff

    return {
        'source': 'csv',
        'metadata': csv_data['metadata'],
        'tracking': tracking,
        'totals': totals,
        'efficiency': efficiency,
        'efficiency_daily': efficiency_daily,
        'daily': daily,
        'work_points': {
            'daily': {wd: csv_data['daily'][wd]['points'] for wd in DAY_ORDER},
            'total': csv_data['totals']['points'],
        },
    }


# ══════════════════════════════════════════════════════════════════════════
#  LEGACY XLSX MODE
# ══════════════════════════════════════════════════════════════════════════

def parse_xlsx(path: str) -> list[dict]:
    """Parse tracker.xlsx and return list of entry dicts."""
    try:
        import openpyxl
    except ImportError:
        print("XLSX mode 需要安裝 openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    entries = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date, weekday, ts, mod, sub, diff, count, mins, rate, note = row
        if date is None or (isinstance(date, str) and '2025' in date):
            continue
        mod_str = str(mod).lower().strip() if mod else ''
        mod_norm = XLSX_MOD_MAP.get(mod_str, mod_str)
        sub_cat = sub or diff or ''
        entries.append({
            'date': date, 'weekday': str(weekday).strip(),
            'time': ts, 'mod': mod_norm, 'sub': str(sub_cat).strip(),
            'count': float(count) if count else 0,
            'mins': float(mins) if mins else 0,
            'rate': float(rate) if rate else 0,
            'note': str(note).strip() if note and str(note) != 'None' else ''
        })
    return entries


def parse_week_input_md(path: str) -> dict:
    """Parse week_input.md for start/end remaining and metadata."""
    text = Path(path).read_text(encoding='utf-8')
    info = {'start': {}, 'end': {}, 'mid': {}, 'study_hr': 0, 'notes': []}

    for line in text.split('\n'):
        line = line.strip()
        if '期初剩餘' in line:
            for mod in ['X光', 'CT', 'US', 'Mammo']:
                m = re.search(rf'{mod}\s*(\d+)', line, re.IGNORECASE)
                if m:
                    info['start'][mod] = int(m.group(1))
        if '期末剩餘' in line:
            for mod in ['X光', 'CT', 'US', 'Mammo']:
                m = re.search(rf'{mod}\s*(\d+)', line, re.IGNORECASE)
                if m:
                    info['end'][mod] = int(m.group(1))
    return info


def group_by_day(entries: list[dict]) -> dict:
    """Group entries by work day (using weekday field)."""
    daily = {}
    for wd in DAY_ORDER:
        daily[wd] = {
            'entries': [], 'report_min': 0, 'clinical_min': 0, 'misc_min': 0,
            'notes': set(),
            **{m: 0 for m in ALL_MODS},
            'xr_sub': defaultdict(lambda: {'c': 0, 'm': 0}),
            'ct_sub': defaultdict(lambda: {'c': 0, 'm': 0}),
            'us_sub': defaultdict(lambda: {'c': 0, 'm': 0}),
        }

    for e in entries:
        wd = e['weekday']
        if wd not in daily:
            continue
        d = daily[wd]
        d['entries'].append(e)
        mod, count, mins = e['mod'], e['count'], e['mins']

        if mod == '其他':
            if '臨床' in e['sub']:
                d['clinical_min'] += mins
            else:
                d['misc_min'] += mins
            continue

        d['report_min'] += mins
        if mod in d:
            d[mod] += count

        if mod == 'X光':
            sub = e['sub'] if e['sub'] else '普通'
            d['xr_sub'][sub]['c'] += count
            d['xr_sub'][sub]['m'] += mins
        elif mod == 'CT':
            sub = e['sub'] if e['sub'] else 'Chest/Abd'
            d['ct_sub'][sub]['c'] += count
            d['ct_sub'][sub]['m'] += mins
        elif mod == 'US':
            sub = e['sub'] if e['sub'] else '一般'
            d['us_sub'][sub]['c'] += count
            d['us_sub'][sub]['m'] += mins

        if e['note']:
            d['notes'].add(e['note'])

    return daily


def compute_totals(daily_or_input: dict, source: str = 'xlsx') -> dict:
    """Compute weekly totals from daily data (xlsx mode)."""
    if source == 'xlsx':
        daily = daily_or_input
        totals = {m: 0 for m in ALL_MODS}
        total_report = total_clinical = total_misc = 0
        for wd in DAY_ORDER:
            d = daily[wd]
            for m in ALL_MODS:
                totals[m] += d[m]
            total_report += d['report_min']
            total_clinical += d['clinical_min']
            total_misc += d['misc_min']
        return {
            'mods': totals,
            'report_min': total_report,
            'clinical_min': total_clinical,
            'misc_min': total_misc,
            'grand_total': sum(totals[m] for m in TRACKED_MODS) + totals.get('MR', 0),
        }
    return daily_or_input  # CSV mode passes pre-computed totals


def compute_efficiency_xlsx(daily: dict) -> dict:
    """Compute efficiency by sub-category across all days (xlsx mode with measured times)."""
    subs = defaultdict(lambda: {'c': 0, 'm': 0})
    for wd in DAY_ORDER:
        d = daily[wd]
        for sub, data in d['xr_sub'].items():
            subs[f'X光({sub})']['c'] += data['c']
            subs[f'X光({sub})']['m'] += data['m']
        for sub, data in d['ct_sub'].items():
            subs[f'CT({sub})']['c'] += data['c']
            subs[f'CT({sub})']['m'] += data['m']
        for sub, data in d['us_sub'].items():
            subs[f'US({sub})']['c'] += data['c']
            subs[f'US({sub})']['m'] += data['m']

    for mod in ['Mammo', 'MR', 'IVP']:
        total_c = sum(daily[wd][mod] for wd in DAY_ORDER)
        total_m = sum(e['mins'] for wd in DAY_ORDER for e in daily[wd]['entries'] if e['mod'] == mod)
        if total_c > 0:
            subs[mod] = {'c': total_c, 'm': total_m}

    result = {}
    for key, data in sorted(subs.items()):
        if data['m'] > 0 and data['c'] > 0:
            actual = data['c'] / (data['m'] / 60)
            baseline = BASELINES.get(key)
            diff_pct = ((actual - baseline) / baseline * 100) if baseline else None
            result[key] = {
                'count': data['c'], 'mins': data['m'],
                'actual_rate': round(actual, 1),
                'baseline': baseline,
                'diff_pct': round(diff_pct, 0) if diff_pct is not None else None
            }
    return result


def compute_tracking(week_input: dict, totals: dict) -> dict:
    """Compute remaining tracking with start/end data."""
    tracking = {}
    for mod in TRACKED_MODS:
        s = week_input['start'].get(mod)
        e = week_input['end'].get(mod)
        completed = totals['mods'].get(mod, 0)
        if s is not None and e is not None:
            added = e - s + completed
            pool = s + added
            digest_rate = completed / pool * 100 if pool > 0 else 0
            tracking[mod] = {
                'start': s, 'end': e, 'completed': completed,
                'added': added, 'net': e - s,
                'digest_rate': round(digest_rate, 1)
            }
        else:
            tracking[mod] = {
                'start': s, 'end': e, 'completed': completed,
                'added': None, 'net': None, 'digest_rate': None
            }
    return tracking


def xlsx_to_report_data(entries: list[dict], week_input: dict) -> dict:
    """Process xlsx entries into standard report data."""
    daily = group_by_day(entries)
    totals = compute_totals(daily)
    efficiency = compute_efficiency_xlsx(daily)
    tracking = compute_tracking(week_input, totals)

    return {
        'source': 'xlsx',
        'tracking': tracking,
        'totals': {k: v if not isinstance(v, float) else round(v, 1)
                   for k, v in totals.items()},
        'efficiency': efficiency,
        'daily': {wd: {
            'total': sum(daily[wd][m] for m in ALL_MODS),
            'report_min': daily[wd]['report_min'],
            'clinical_min': daily[wd]['clinical_min'],
            **{m: daily[wd][m] for m in ALL_MODS if daily[wd][m] > 0}
        } for wd in DAY_ORDER},
    }


# ══════════════════════════════════════════════════════════════════════════
#  OUTPUT
# ══════════════════════════════════════════════════════════════════════════

def print_report(report_data: dict):
    """Print human-readable summary to stdout."""
    source = report_data.get('source', 'unknown')
    tracking = report_data['tracking']
    totals = report_data['totals']
    daily = report_data['daily']
    efficiency = report_data.get('efficiency', {})

    print("=" * 70)
    print(f"WEEKLY REPORT DATA (source: {source})")
    print("=" * 70)

    print("\n── Tracking ──")
    for mod, t in tracking.items():
        if t['net'] is not None:
            print(f"  {mod}: {t['start']} -> {t['end']} (net {t['net']:+d}, "
                  f"completed {t['completed']:.0f}, added >={t['added']:.0f}, "
                  f"digest {t['digest_rate']}%)")

    print(f"\n── Totals ──")
    mods = totals.get('mods', {})
    for mod in ALL_MODS:
        if mods.get(mod, 0) > 0:
            print(f"  {mod}: {mods[mod]:.0f}")
    print(f"  Grand total: {totals.get('grand_total', 0):.0f}")

    if source == 'csv':
        active_hr = totals.get('active_hr', 0)
        print(f"  Active time (estimated): {active_hr:.1f}hr")
        print(f"  Work points: {totals.get('points', 0):.1f}")
    else:
        report_min = totals.get('report_min', 0)
        print(f"  Report time: {report_min:.0f}m ({report_min/60:.1f}hr)")
        clin = totals.get('clinical_min', 0)
        misc = totals.get('misc_min', 0)
        print(f"  Clinical: {clin:.0f}m, Misc: {misc:.0f}m")

    print(f"\n── Daily ──")
    for wd in DAY_ORDER:
        d = daily[wd]
        day_total = d.get('total', d.get('cases', 0))
        if day_total > 0:
            if source == 'csv':
                active = d.get('active_hr', 0)
                pts = d.get('points', 0)
                print(f"  {wd}: {day_total:.0f}案 / ~{active:.1f}hr / {pts:.1f}pt")
            else:
                report_min = d.get('report_min', 0)
                clin = f" +clin {d.get('clinical_min', 0):.0f}m" if d.get('clinical_min', 0) > 0 else ""
                print(f"  {wd}: {day_total:.0f}份 / {report_min:.0f}m ({report_min/60:.1f}hr){clin}")

    if efficiency:
        print(f"\n── Efficiency ──")
        for key, e in sorted(efficiency.items()):
            if e.get('estimated'):
                print(f"  {key}: {e['count']}案 [base: {e.get('baseline')}] (時間無法從CSV推算)")
            elif e.get('mins') and e.get('actual_rate'):
                diff_str = f" ({e['diff_pct']:+.0f}%)" if e.get('diff_pct') is not None else ""
                print(f"  {key}: {e['count']:.0f}份/{e['mins']:.0f}m = {e['actual_rate']}/hr"
                      f" [base: {e['baseline']}]{diff_str}")

    if source == 'csv':
        eff_daily = report_data.get('efficiency_daily', {})
        if eff_daily:
            print(f"\n── Daily Efficiency (estimated) ──")
            for wd in DAY_ORDER:
                if wd in eff_daily:
                    ed = eff_daily[wd]
                    print(f"  {wd}: {ed['cases']}cases / {ed['active_hr']}hr "
                          f"~= {ed['rate']}cases/hr (est.)")

    print(f"\n── Output ──")
    print(f"  Report data ready. Use Claude to generate HTML or pipe to template.")


# ══════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Generate radiology weekly report',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  CSV mode:   python generate_report.py --csv 202602.csv --input week_input.yaml
  XLSX mode:  python generate_report.py --xlsx tracker.xlsx --input week_input.md
  With YK:    python generate_report.py --csv 202602.csv --yk 202602YK.csv --input week_input.yaml
        """)

    # Input mode (mutually exclusive)
    mode = parser.add_mutually_exclusive_group(required=True)
    mode.add_argument('--csv', nargs='+', help='CSV file(s) from hospital system')
    mode.add_argument('--xlsx', help='Legacy tracker.xlsx')

    parser.add_argument('--yk', help='YK campus duty CSV (optional, CSV mode only)')
    parser.add_argument('--input', required=True, help='week_input.yaml or week_input.md')
    parser.add_argument('--reporter', default='A80748', help='Reporter ID (CSV mode)')
    parser.add_argument('--mid', action='store_true',
                        help='Mid-week mode: use remaining.mid instead of remaining.end')
    parser.add_argument('-o', '--output', default='output/weekly_report.json',
                        help='Output JSON path')

    args = parser.parse_args()

    if args.csv:
        # ── CSV Mode ──
        input_path = args.input
        if input_path.endswith('.yaml') or input_path.endswith('.yml'):
            week_input = parse_yaml_input(input_path)
        else:
            week_input = parse_week_input_md(input_path)

        # Mid-week mode: swap mid -> end
        if args.mid:
            mid = week_input.get('mid', {})
            if mid:
                week_input['end'] = mid
                print("Mid-week mode: using remaining.mid as end values", file=sys.stderr)
            else:
                print("Warning: --mid specified but no remaining.mid in week_input.yaml",
                      file=sys.stderr)
            week_input['mode'] = 'midweek'

        week_str = week_input.get('week', '')
        if not week_str:
            print("Error: week_input must contain 'week' field (e.g., '2026-W09')",
                  file=sys.stderr)
            sys.exit(1)

        csv_data = load_csv_data(args.csv, args.yk, week_str, args.reporter)
        report_data = csv_to_report_data(csv_data, week_input)

        # Tag midweek mode in output
        if args.mid:
            report_data['mode'] = 'midweek'
    else:
        # ── Legacy XLSX Mode ──
        entries = parse_xlsx(args.xlsx)
        week_input = parse_week_input_md(args.input)
        report_data = xlsx_to_report_data(entries, week_input)

    # Print summary
    print_report(report_data)

    # Save JSON
    json_path = args.output
    if json_path.endswith('.html'):
        json_path = json_path.replace('.html', '.json')
    Path(json_path).parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(report_data, f, ensure_ascii=False, indent=2, default=str)
    print(f"  JSON saved to: {json_path}")


if __name__ == '__main__':
    main()
